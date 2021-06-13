from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from datetime import datetime
from gettingMethod.GetInformation import sendingCommonRequest
import blackduck

thin = Side(border_style="thin", color="000000")


# Write Project Name
def writeProjectName(ws, prjName, prjVersion):
    ws['a1'].value = f'Project: {prjName} - {prjVersion}'
    ws['a1'].font = Font(bold=True)


# Write Today
def writeToday(ws, start_row, end_row):
    ws.merge_cells(f'{start_row}:{end_row}')
    ws[start_row].value = datetime.now()
    ws[start_row].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


# Creating Default Form of the Report
def createDefaultForm(ws, start_row, end_row, value, PattenColor="d9d9d9", fontColor="000000"):
    ws.merge_cells(f'{start_row}:{end_row}')
    ws[start_row].value = value
    ws[start_row].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[start_row].fill = PatternFill("solid", fgColor=PattenColor)
    ws[start_row].font = Font(bold=True, color=fontColor)


# Write One Value
def writeDefaultValue(ws, cell, value="", PattenColor="FFFFFF", fontColor="000000"):
    ws[cell].value = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell].fill = PatternFill("solid", fgColor=PattenColor)
    ws[cell].font = Font(bold=True, color=fontColor)


# Write Data
def writeData(ws, componentInfo, rowNum):
    # 4 행부터 입력 시작
    ws[f'a{rowNum}'].value = ' ' + componentInfo['componentName']
    try:
        ws[f'b{rowNum}'].value = componentInfo['componentVersionName']
    except KeyError:
        ws[f'b{rowNum}'].value = 'Unknown'

    ws[f'c{rowNum}'].value = ' ' + componentInfo['licenses'][0]['licenseDisplay']

    matchType = componentInfo['matchTypes'][0]

    # Match Type이 여러개일 경우 합치기
    matchTypeLen = len(componentInfo['matchTypes'])
    if matchTypeLen > 1:
        for i in range(1, matchTypeLen):
            matchType = matchType + "\n" + componentInfo['matchTypes'][i]
    ws[f'd{rowNum}'].value = matchType

    # Critical, High, Medium 취약점 수
    ws[f'e{rowNum}'].value = componentInfo['securityRiskProfile']['counts'][5]['count']
    ws[f'f{rowNum}'].value = componentInfo['securityRiskProfile']['counts'][4]['count']
    ws[f'g{rowNum}'].value = componentInfo['securityRiskProfile']['counts'][3]['count']

    # Component Upgrade Guide 정보 리턴
    if componentInfo['_meta']['links'][7] in componentInfo['_meta']['links']:
        guidanceInfo = sendingCommonRequest(componentInfo['_meta']['links'][7]['href'], blackduck.bearerToken).json()

        # ShortTerm Guidance
        try:
            ws[f'h{rowNum}'].value = guidanceInfo['shortTerm']['versionName']
            ws[f'i{rowNum}'].value = guidanceInfo['shortTerm']['vulnerabilityRisk']['critical'] \
                                     + guidanceInfo['shortTerm']['vulnerabilityRisk']['high'] \
                                     + guidanceInfo['shortTerm']['vulnerabilityRisk']['medium']
        except KeyError:
            ws[f'h{rowNum}'].value = 'N/A'
            ws[f'i{rowNum}'].value = 'N/A'

        # LongTerm Guidance
        try:
            ws[f'j{rowNum}'].value = guidanceInfo['longTerm']['versionName']
            ws[f'k{rowNum}'].value = guidanceInfo['longTerm']['vulnerabilityRisk']['critical'] \
                                     + guidanceInfo['longTerm']['vulnerabilityRisk']['high'] \
                                     + guidanceInfo['longTerm']['vulnerabilityRisk']['medium']
        except KeyError:
            ws[f'j{rowNum}'].value = 'N/A'
            ws[f'k{rowNum}'].value = 'N/A'
    else:
        ws[f'h{rowNum}'].value = 'N/A'
        ws[f'i{rowNum}'].value = 'N/A'
        ws[f'j{rowNum}'].value = 'N/A'
        ws[f'k{rowNum}'].value = 'N/A'

    ws[f'a{rowNum}'].alignment = Alignment(horizontal="left", vertical="center", wrapText=True)
    ws[f'b{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'c{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'd{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'e{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'f{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'g{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'h{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'i{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'j{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws[f'k{rowNum}'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


def drawBorders(ws, end_row):
    clmnList = 'ABCDEFGHIJK'
    for num in range(2, end_row):
        for char in clmnList:
            ws[f'{char}{num}'].border = Border(right=thin, left=thin, top=thin, bottom=thin)
